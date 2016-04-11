from openpyxl import Workbook, load_workbook
import os, ctypes, getpass, easygui

print'---------------------------------------------------------------------------'
print '                      COMPARE LISTS APPLICATION'
print'---------------------------------------------------------------------------'
print
#Select First Document
while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open the first .xlsx list', 'Open', 1)
	list_one = easygui.fileopenbox()
	try:
		first_wb = load_workbook(filename = list_one, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid name', 'Import Error!', 1)
		continue
active_sheet_1 = first_wb.get_sheet_names()[0]
working_sheet_1 = first_wb.get_sheet_by_name(active_sheet_1) 
first_list_description = raw_input('Enter the description of the first list\n')

#Select Second Document
while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open the second .xlsx list', 'Open', 1)
	list_two = easygui.fileopenbox()
	try:
		second_wb = load_workbook(filename = list_two, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid name', 'Import Error!', 1)
		continue
active_sheet_2 = second_wb.get_sheet_names()[0]
working_sheet_2 = second_wb.get_sheet_by_name(active_sheet_2) 
second_list_description = raw_input('Enter the description of the second list\n')

#Create txt document to write results

results_file = 'Comparison_Results.txt'
results_fhand = open(results_file, 'w')

#Run the list comparisons and write the results to the txt file
print 'Checking for devices found in the ' + second_list_description + ' list that were not found in ' + first_list_description + ' list ...'
results_fhand.write('Checking for devices found in the ' + second_list_description + ' that were not found in ' + first_list_description+ ' list...\n\n')

check_list_1 = list()
for row in working_sheet_1:
	for cell in row:
		check_list_1.append(cell.value.upper())

count_1 = 0		
for row in working_sheet_2:
	for cell in row:
		if cell.value.upper() not in check_list_1:
			print cell.value
			results_fhand.write(cell.value + '\n')
			count_1 += 1
		else:
			continue

print 'Number of Devices from the ' + second_list_description+ ' list that are not in ' + first_list_description+ ' list: ', count_1
results_fhand.write('Number of Devices from the ' + second_list_description + ' list that are not in ' + first_list_description + ' list: ' + str(count_1) + '\n\n')
print

print 'Checking for devices found in the ' + first_list_description + ' list that were not found in the ' + second_list_description + ' list...'
results_fhand.write('Checking for devices found in the ' + first_list_description + ' list that were not found in the ' + second_list_description + ' list ...\n\n')

check_list_2 = list()
for row in working_sheet_2:
	for cell in row:
		check_list_2.append(cell.value.upper())

count_2 = 0		
for row in working_sheet_1:
	for cell in row:
		if cell.value.upper() not in check_list_2:
			print cell.value
			results_fhand.write(cell.value + '\n')
			count_2 += 1
		else:
			continue			
print 'Number of devices from the ' + first_list_description + ' list that are not in the ' + second_list_description + ' list: ', count_2
results_fhand.write('Number of Devices from the ' + second_list_description + ' list that are not in the ' + first_list_description + ' list: ' + str(count_2) + '\n\n')
print
print 'Application Complete!'
