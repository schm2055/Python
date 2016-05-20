from openpyxl import Workbook, load_workbook
import os, ctypes, getpass, easygui
import sqlite3

#Connect to SQL DB - creates DB in the directory if it does not exist at the time of execution
conn_db = sqlite3.connect('appmapping.sqlite')
cur_db = conn_db.cursor()
print 'Database created and/or connected to successfully!'

#Create SQL tables - One for Services and one for Service Instances
cur_db.executescript('''

CREATE TABLE IF NOT EXISTS Services(
	id TEXT,
	Display_Name TEXT,
	Name TEXT,
	Service_Type TEXT,
	Description TEXT,
	Vendor TEXT,
	Category TEXT,
	Notes TEXT,
	Application TEXT
);

CREATE TABLE IF NOT EXISTS Service_Instances(
	id TEXT,
	Service TEXT,
	Start_Mode TEXT,
	State TEXT,
	Dont_change_via_api TEXT,
	Device TEXT,
	User TEXT,
	First_Detected TEXT,
	Last_Updated TEXT,
	Application_Services TEXT
);
''')

print 'Database Tables Created/Verified Successfully!'

#Import the Services spreadsheet exported from Device42
while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open your Services Document', 'Open', 1)
	services = easygui.fileopenbox()
	try:
		services_wb = load_workbook(filename = services, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid services inventory  name', 'Import Error!', 1)
		continue
services_active_sheet = services_wb.get_sheet_names()[0]
services_working_sheet = services_wb.get_sheet_by_name(services_active_sheet) 

#Import he Service Instances spreadsheet exported from Device 42
while True:
	ctypes.windll.user32.MessageBoxA(0, 'Open your Service Instances Document', 'Open', 1)
	instances = easygui.fileopenbox()
	try:
		instances_wb = load_workbook(filename = instances, use_iterators = True)
		break
	except:
		ctypes.windll.user32.MessageBoxA(0, 'Please enter a valid Service Instances inventory name', 'Import Error!', 1)
		continue
instances_active_sheet = instances_wb.get_sheet_names()[0]
instances_working_sheet = instances_wb.get_sheet_by_name(instances_active_sheet)

#Define function to get row values in a list
def iter_rows(working_sheet):
    for row in working_sheet.iter_rows():
        yield [cell.value for cell in row]

#Import Services Contents into Services Table		
services_row_list = iter_rows(services_working_sheet)

for list in services_row_list:
	id = list[0]
	display_name = list[1]
	name = list[2]
	service_type = list[3]
	description = list[4]
	vendor = list[5]
	category = list[6]
	notes = list[7]
	application = list[8]
	cur_db.execute('''INSERT INTO Services (id, Display_Name, Name, Service_Type, Description, Vendor, Category, Notes, Application)
		VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (id, display_name, name, service_type, description, vendor, category, notes, application, ))
	conn_db.commit()

print 'Services Imported Successfully!'
	
#Import Service Instances contents into the Service Instances table
instances_row_list = iter_rows(instances_working_sheet)

for list in instances_row_list:
	id = list[0]
	service = list[1]
	start_mode = list[2]
	state = list[3]
	dont_exchange_via_api = list[4]
	device = list[5]
	user = list[6]
	first_detected = list[7]
	last_updated = list[8]
	application_services = list[9]
	cur_db.execute('''INSERT INTO Service_Instances (id, Service, Start_Mode, State, Dont_Change_via_api, Device, User, First_Detected, Last_Updated, Application_Services)
		VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (id, service, start_mode, state, dont_exchange_via_api, device, user, first_detected,last_updated, application_services, ))
	conn_db.commit()
	
print 'Service Instances Imported Successfully!'

#Populate application services field on Service Instances table
cur_db.execute('''SELECT Service FROM Service_Instances''')
service_list = [record[0] for record in cur_db.fetchall()]

for service in service_list:
	cur_db.execute("SELECT Application FROM Services WHERE Display_Name = ?", (service, ))
	app_list = [record[0] for record in cur_db.fetchall()]
	try:
		app_val = app_list[0]
	except:
		continue
	cur_db.execute("UPDATE Service_Instances SET Application_Services = ? WHERE Service = ?", (app_val, service, ))
	conn_db.commit()

